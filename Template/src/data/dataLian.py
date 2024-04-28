from openpyxl import load_workbook
from src.player.playerLian import *

class data:
    FIRST_ROW = 8
    FIRST_COLONNE = 1
    LAST_COLONNE = 5

    FIRST_COLONNE_ROUGE = 1
    FIRST_COLONNE_BLEU = 8


    def __init__(self, Path, equipeBleu,equipeRouge):
        self.equipeBleu = equipeBleu
        self.equipeRouge = equipeRouge

        self.path = Path
        # self.wb = load_workbook(filename=self.path, read_only=True)
        # self.sheet = self.wb.active  # Ou self.wb['NomDeLaFeuille']
        
        self.cheminExcel = r"C:\Docker\devenvironments\repository\Template\nomenclature.xlsx"
        self.fichierExcel = load_workbook(filename=self.cheminExcel, read_only=True,data_only=True)
        self.feuille = self.fichierExcel.active

        # Lire cellule -> Nombre de VM pour chaque
        self.nbRouge = self.feuille['B2'].value
        self.nbBleu = self.feuille['B3'].value
  
        self.extraireDonnees()
    
    def obtenirLigne (self,index, debutColonne):
        return [self.feuille.cell(row=index+self.FIRST_ROW, column=debutColonne + i).value for i in range(self.LAST_COLONNE)]
    
    def obtenirUserCible (self,index):
        return self.feuille.cell(row=index+self.FIRST_ROW, column=14).value
    
    def extraireDonnees(self): 
        # tableau = self.feuille.iter_rows(min_row=data.FIRST_ROW, min_col=(data.FIRST_COLONNE + debutColonne), max_col=(data.LAST_COLONNE + debutColonne),max_row=(nbVM+data.FIRST_ROW))
        for index in range(20):
            listeBleu = self.obtenirLigne(index,self.FIRST_COLONNE_BLEU)
            listeRouge = self.obtenirLigne(index,self.FIRST_COLONNE_ROUGE)
            userCible = self.obtenirUserCible(index)
            
            joueurBleu = playerVictime(listeBleu,self.equipeBleu,userCible)
            self.equipeBleu.addPlayer(joueurBleu)

            joueurRouge = playerAttaque(listeRouge,self.equipeRouge,joueurBleu)
            self.equipeRouge.addPlayer(joueurRouge)
        self.fichierExcel.close()

# Fonctions appelees par le main
    def get_nb_bleu(self):
        return self.nbBleu

    def get_nb_rouge(self):
        return self.nbRouge
